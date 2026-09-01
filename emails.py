"""HPCL LIVEDB — Exception mail notifications via Microsoft Outlook COM.

Emails are sent through the Outlook desktop application running on the same
Windows machine as this Streamlit app.  No SMTP credentials needed —
Outlook uses the currently signed-in HPCL account.

Requirements (local only):
  • Microsoft Outlook must be installed and OPEN on this PC
  • shoaibrehman@hpcl.in must be the active Outlook account
  • pywin32 installed:  pip install pywin32

On Streamlit Cloud (Linux), Outlook COM is unavailable — the UI shows a
notice and disables the Send button automatically.
"""

from __future__ import annotations

import base64
import gc
import io
import os
import platform
import re
import tempfile
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd

# ── Logo loader (embed HPCL logo as base64 data-URI for email) ─────────────────
_BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
_LOGO_PATH  = os.path.join(_BASE_DIR, "MAster", "Master Logo.jpg")

def _logo_data_uri() -> str:
    """Return base64-encoded data URI for HPCL Master Logo, or empty string."""
    try:
        if os.path.exists(_LOGO_PATH):
            with open(_LOGO_PATH, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()
            return f"data:image/jpeg;base64,{b64}"
    except Exception:
        pass
    return ""

# ── Constants ─────────────────────────────────────────────────────────────────

SENDER_EMAIL = "shoaibrehman@hpcl.in"

BCC_EMAILS: List[str] = [
    "SOD.OPNS.HQO@hpcl.in",
    "shubham.tayal@hpcl.in",
]

# ── Zone → recipient map ──────────────────────────────────────────────────────

ZONE_EMAIL_MAP: Dict[str, Dict[str, str]] = {
    "Bengaluru Zone":       {"to": "brijeshkumar@hpcl.in",              "cc": "BLR.OND.IC@hpcl.in"},
    "Bhopal Zone":          {"to": "agajare@hpcl.in;twinacore@hpcl.in", "cc": "CZ.OND.IC@hpcl.in"},
    "Bhubaneshwar Zone":    {"to": "smarak.lenka@hpcl.in",              "cc": "ECZ.OND.IC@hpcl.in"},
    "Chandigarh Zone":      {"to": "haroonhamid@hpcl.in",               "cc": "NFZ.OND.IC@hpcl.in"},
    "Cochin Zone":          {"to": "kathir@hpcl.in",                    "cc": "kbanothu@hpcl.in"},
    "East Zone":            {"to": "sray@hpcl.in",                      "cc": "EZ.OND.IC@hpcl.in"},
    "Guwahati Zone":        {"to": "lodyuo@hpcl.in",                    "cc": "gurubachansingha@hpcl.in"},
    "Jaipur Zone":          {"to": "rjprasad@hpcl.in",                  "cc": "NWF.OND.IC@hpcl.in"},
    "Noida (UP-West) Zone": {"to": "aradhnat@hpcl.in",                  "cc": "chraghu@hpcl.in"},
    "North Central Zone":   {"to": "rvpandey@hpcl.in",                  "cc": "adeshmukh@hpcl.in"},
    "North West Zone":      {"to": "sanjaykdewangan@hpcl.in",           "cc": "NWZ.OND.IC@hpcl.in"},
    "North Zone (NZ)":      {"to": "ajaygr@hpcl.in",                    "cc": "NZ.OND.IC@hpcl.in"},
    "Patna Zone":           {"to": "ajaisingh@hpcl.in",                 "cc": "dastidar@hpcl.in"},
    "South Central Zone":   {"to": "sangamkmohan@hpcl.in",              "cc": "SCRZ.OND.IC@hpcl.in"},
    "South Zone":           {"to": "venkates@hpcl.in",                  "cc": "SZ.OND.IC@hpcl.in"},
    "West Zone (WZ)":       {"to": "ntgajbiye@hpcl.in",                 "cc": "WZ.OND.IC@hpcl.in"},
}

# ── Exception labels (key = column name in all_exception_plant_df) ────────────

EXCEPTION_LABELS: Dict[str, str] = {
    "Pending DC":                    "Pending DCs",
    "Open Delivery":                 "Open Deliveries",
    "Open In-Transit":               "Open In-Transit STOs",
    "Open Sales Order":              "Open Sales Orders",
    "Pending Invoice":               "Pending Invoices",
    "Shortage Sales (Billing Docs)": "Shortages - Sales",
    "Shortage STO (Billing Docs)":   "Shortages - STO",
}

# ── Outlook availability check ────────────────────────────────────────────────

def outlook_available() -> tuple[bool, str]:
    """Return (True, '') if Outlook COM can be used; else (False, reason)."""
    if platform.system() != "Windows":
        return False, "Outlook COM is only available on Windows (not on Streamlit Cloud)."
    try:
        import pythoncom  # noqa: F401
        import win32com.client  # noqa: F401
        return True, ""
    except ImportError:
        return False, "pywin32 not installed. Run: pip install pywin32"


# ── Safe filename helper ──────────────────────────────────────────────────────

def _safe_name(s: str) -> str:
    """Strip characters not allowed in Windows filenames."""
    return re.sub(r'[\\/:*?"<>|]', '', s).strip()


# ── Low-level Outlook sender ──────────────────────────────────────────────────

def _send_via_outlook(
    to: str,
    subject: str,
    html_body: str,
    cc: str = "",
    bcc: str = "",
    attachments: Optional[List[Tuple[str, bytes]]] = None,
) -> dict:
    """Send one email through Outlook COM.

    attachments: list of (display_filename, bytes) tuples
    Returns {"ok": True/False, "mode": "sent"/"draft", "msg": "..."}
    """
    import pythoncom
    import win32com.client as win32

    com_init = False
    outlook = mail_item = None
    tmp_paths: List[str] = []

    try:
        pythoncom.CoInitialize()
        com_init = True

        outlook = win32.Dispatch("Outlook.Application")
        mail_item = outlook.CreateItem(0)
        mail_item.To      = to
        mail_item.Subject = subject
        mail_item.HTMLBody = html_body
        if cc.strip():
            mail_item.CC = cc
        if bcc.strip():
            mail_item.BCC = bcc

        if attachments:
            for fname, fbytes in attachments:
                ext = os.path.splitext(fname)[1] or ".xlsx"
                tmp = tempfile.NamedTemporaryFile(
                    delete=False, suffix=ext, prefix="hpcl_"
                )
                tmp.write(fbytes)
                tmp.close()
                tmp_paths.append((tmp.name, fname))

            for tmp_path, display_name in tmp_paths:
                mail_item.Attachments.Add(tmp_path, 1, 1, display_name)

        try:
            mail_item.Send()
            return {"ok": True, "mode": "sent", "msg": "Mail sent successfully."}
        except Exception as send_exc:
            try:
                mail_item.Save()
                return {"ok": True, "mode": "draft",
                        "msg": f"Saved to Drafts (Send failed: {send_exc})"}
            except Exception as draft_exc:
                return {"ok": False,
                        "msg": f"Send and draft both failed: {draft_exc}"}

    except Exception as exc:
        return {"ok": False, "msg": f"Outlook error: {exc}"}

    finally:
        try:
            del mail_item
        except Exception:
            pass
        try:
            del outlook
        except Exception:
            pass
        gc.collect()
        if com_init:
            try:
                import pythoncom as _pc
                _pc.CoUninitialize()
            except Exception:
                pass
        for tmp_path, _ in tmp_paths:
            try:
                os.remove(tmp_path)
            except Exception:
                pass


# ── Excel builder — one file per exception type ───────────────────────────────

def build_zone_excel_attachments(
    zone_name: str,
    detail_dfs: Dict[str, pd.DataFrame],
    selected_exceptions: List[str],
) -> List[Tuple[str, bytes]]:
    """Build one Excel file per selected exception type for a zone.

    Returns list of (filename, bytes) tuples ready to attach.
    Filename format:  ZoneName_ExceptionLabel.xlsx
    Only exception types that have rows for this zone are included.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    results: List[Tuple[str, bytes]] = []

    for exc_key in selected_exceptions:
        label = EXCEPTION_LABELS.get(exc_key, exc_key)
        detail_df = detail_dfs.get(exc_key)

        if detail_df is None or not isinstance(detail_df, pd.DataFrame) or detail_df.empty:
            continue

        # Filter to this zone
        if "Zone Name" in detail_df.columns:
            zone_df = detail_df[detail_df["Zone Name"] == zone_name].copy()
        else:
            zone_df = detail_df.copy()  # can't filter — include all

        if zone_df.empty:
            continue

        # Drop purely internal helper columns
        drop_cols = [c for c in zone_df.columns
                     if c.startswith("_") or c in {"index", "level_0"}]
        zone_df = zone_df.drop(columns=drop_cols, errors="ignore").reset_index(drop=True)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            zone_df.to_excel(writer, index=False, sheet_name=label[:31])
            ws = writer.sheets[label[:31]]

            hdr_fill = PatternFill("solid", fgColor="003087")
            hdr_font = Font(bold=True, color="FFFFFF", size=10)
            thin      = Side(style="thin", color="D5E2F3")
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            center    = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cell in ws[1]:
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = center
                cell.border    = border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = border

            for col in ws.columns:
                max_len = max(
                    (len(str(c.value or "")) for c in col),
                    default=10
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

            ws.row_dimensions[1].height = 28

        safe_zone  = _safe_name(zone_name)
        safe_label = _safe_name(label)
        filename   = f"{safe_zone}_{safe_label}.xlsx"
        results.append((filename, buf.getvalue()))

    return results


# ── HTML email builder ────────────────────────────────────────────────────────

def _kpi_tile_html(label: str, value: str, sub: str, border_color: str, bg: str) -> str:
    """Render one KPI tile cell (email-safe, table-based, no flex/grid)."""
    return (
        f'<td width="25%" style="padding:4px;">'
        f'<table width="100%" cellpadding="0" cellspacing="0">'
        f'<tr><td style="background:{bg};border-top:4px solid {border_color};'
        f'border-radius:8px;padding:12px 10px 10px;text-align:center;">'
        f'<div style="font-size:9px;font-weight:700;color:#666;text-transform:uppercase;'
        f'letter-spacing:.07em;margin-bottom:5px;">{label}</div>'
        f'<div style="font-size:20px;font-weight:900;color:{border_color};'
        f'line-height:1.1;">{value}</div>'
        f'<div style="font-size:9px;color:#888;margin-top:3px;">{sub}</div>'
        f'</td></tr></table>'
        f'</td>'
    )


def build_exception_email_html(
    zone_name: str,
    as_of_date: str,
    exception_summary_df: pd.DataFrame,
    selected_exceptions: List[str],
    custom_intro: str = "",
    attachment_names: Optional[List[str]] = None,
    zone_kpi_dict: Optional[dict] = None,
) -> str:
    """Build HPCL-branded HTML email body for one zone's exceptions.

    zone_kpi_dict: pre-computed per-zone KPI values dict (from _build_zone_kpi_dict).
    """

    primary   = "#003087"
    secondary = "#0057A8"
    accent    = "#003087"   # deep HPCL blue (used for company name text)
    _logo_uri = _logo_data_uri()
    as_of     = as_of_date or datetime.now().strftime("%d %b %Y")
    kpis      = zone_kpi_dict or {}

    # Only include columns that exist and are selected
    exc_cols = [c for c in selected_exceptions if c in exception_summary_df.columns]

    if exc_cols:
        mask = (
            exception_summary_df[exc_cols]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .sum(axis=1) > 0
        )
        zone_df = exception_summary_df[mask].copy()
    else:
        zone_df = exception_summary_df.copy()

    total_locations = len(zone_df)

    # ── KPI Tiles block ────────────────────────────────────────────────────────
    def _n(key, default=0):
        return kpis.get(key, default)

    def _tile_row(tiles):
        cells = "".join(
            _kpi_tile_html(t["label"], t["value"], t["sub"], t["color"], t["bg"])
            for t in tiles
        )
        return f'<tr>{cells}</tr>'

    def _fmt_count(v):
        try:
            return f"{int(v):,}"
        except Exception:
            return str(v)

    def _fmt_ltrs(v):
        try:
            n = float(v)
            if n >= 1_000_000:
                return f"{n/1_000_000:.2f}M L"
            if n >= 1_000:
                return f"{n/1_000:.1f}K L"
            return f"{n:,.0f} L"
        except Exception:
            return str(v)

    row1_tiles = [
        {"label": "Pending DCs",        "value": _fmt_count(_n("Pending DC")),
         "sub": "Unique shipments",     "color": "#CC2929", "bg": "#FFF0F0"},
        {"label": "Open Deliveries",    "value": _fmt_count(_n("Open Delivery")),
         "sub": "Unique deliveries",    "color": "#CC2929", "bg": "#FFF0F0"},
        {"label": "Open In-Transit",    "value": _fmt_count(_n("Open In-Transit")),
         "sub": "Unique STO orders",    "color": "#0369A1", "bg": "#EFF8FF"},
        {"label": "Open Sales Orders",  "value": _fmt_count(_n("Open Sales Order")),
         "sub": "Unique sales docs",    "color": "#D97706", "bg": "#FFFBEB"},
    ]
    row2_tiles = [
        {"label": "Pending Invoices",   "value": _fmt_count(_n("Pending Invoice")),
         "sub": "Unique deliveries",    "color": "#D97706", "bg": "#FFFBEB"},
        {"label": "Shortage Sales",     "value": _fmt_ltrs(_n("Shortage Sales (Ltrs)")),
         "sub": "Total Litres",         "color": "#CC2929", "bg": "#FFF0F0"},
        {"label": "Shortage STO",       "value": _fmt_ltrs(_n("Shortage STO (Ltrs)")),
         "sub": "Total Litres",         "color": "#CC2929", "bg": "#FFF0F0"},
        {"label": "Tank Reco",          "value": _fmt_count(_n("Tank Reco")),
         "sub": "Plant+Tank+Material",  "color": "#7C3AED", "bg": "#F5F3FF"},
    ]
    row3_tiles = [
        {"label": "PL Unblock Qty",
         "value": f"{_n('PL Unblock (KL)', 0.0):.3f} KL",
         "sub": "Pipeline stock",       "color": "#0369A1", "bg": "#EFF8FF"},
        {"label": "Dummy Tank Qty",
         "value": f"{_n('Dummy Tank (KL)', 0.0):.3f} KL",
         "sub": "Excl. DBIT/DLUB/DSLP","color": "#D97706", "bg": "#FFFBEB"},
        {"label": "Tank Turns",
         "value": f"{_n('Tank Turns', 0.0):.2f}",
         "sub": "Dispatches÷Capacity",  "color": "#15803D", "bg": "#F0FDF4"},
        {"label": "Location Visit",
         "value": f"{int(_n('Locations Visited', 0))} loc | {_n('Location Compliance (%)', 0.0):.1f}%",
         "sub": "Visited | Compliance", "color": "#15803D", "bg": "#F0FDF4"},
    ]

    kpi_tiles_html = (
        '<table width="100%" cellpadding="0" cellspacing="0">'
        + _tile_row(row1_tiles)
        + '<tr><td colspan="4" style="height:6px;"></td></tr>'
        + _tile_row(row2_tiles)
        + '<tr><td colspan="4" style="height:6px;"></td></tr>'
        + _tile_row(row3_tiles)
        + '</table>'
    ) if kpis else ""

    # ── Location-wise exception table ──────────────────────────────────────────
    col_heads = "".join(
        f'<th style="background:{secondary};color:#fff;padding:7px 8px;'
        f'font-size:10px;font-weight:700;text-align:center;'
        f'border:1px solid rgba(255,255,255,0.2);white-space:nowrap;">'
        f'{EXCEPTION_LABELS.get(c, c)}</th>'
        for c in exc_cols
    )
    table_header = (
        f'<tr>'
        f'<th style="background:{primary};color:{accent};padding:7px 10px;'
        f'font-size:10px;font-weight:700;text-align:left;'
        f'border:1px solid rgba(255,255,255,0.2);">Location</th>'
        f'{col_heads}'
        f'<th style="background:{primary};color:{accent};padding:7px 8px;'
        f'font-size:10px;font-weight:700;text-align:center;'
        f'border:1px solid rgba(255,255,255,0.2);">Total</th>'
        f'</tr>'
    )

    def _fmt_cell(v):
        try:
            n = int(float(v))
            color  = "#CC0000" if n > 0 else "#007700"
            weight = "700" if n > 0 else "400"
            return f'<span style="color:{color};font-weight:{weight};">{n:,}</span>'
        except Exception:
            return str(v)

    rows_html = ""
    for i, (_, row) in enumerate(zone_df.iterrows()):
        bg    = "#FFFFFF" if i % 2 == 0 else "#F4F8FF"
        cells = "".join(
            f'<td style="padding:6px 8px;font-size:10px;text-align:center;'
            f'border:1px solid #E2EAF4;background:{bg};">'
            f'{_fmt_cell(row.get(c, 0))}</td>'
            for c in exc_cols
        )
        total_val  = int(pd.to_numeric(row.get("Total Exceptions", 0), errors="coerce") or 0)
        plant_name = row.get("Plant Name", row.get("Location", ""))
        rows_html += (
            f'<tr>'
            f'<td style="padding:6px 10px;font-size:10px;font-weight:600;'
            f'color:{primary};border:1px solid #E2EAF4;background:{bg};">'
            f'{plant_name}</td>'
            f'{cells}'
            f'<td style="padding:6px 8px;font-size:10px;font-weight:700;'
            f'text-align:center;border:1px solid #E2EAF4;background:{bg};'
            f'color:#CC0000;">{total_val:,}</td>'
            f'</tr>'
        )

    if not rows_html:
        rows_html = (
            f'<tr><td colspan="{len(exc_cols)+2}" style="text-align:center;'
            f'padding:16px;color:#666;font-size:11px;">'
            f'No location-level exceptions found for this zone.</td></tr>'
        )

    # ── Attachments list ───────────────────────────────────────────────────────
    if attachment_names:
        attach_items = "".join(
            f'<li style="margin:3px 0;font-size:11px;">{n}</li>'
            for n in attachment_names
        )
        attach_block = (
            f'<tr><td style="padding:10px 28px 16px;">'
            f'<p style="margin:0 0 5px;font-size:12px;font-weight:700;color:{primary};">'
            f'&#128206; Attachments:</p>'
            f'<ul style="margin:0;padding-left:18px;color:#333;">{attach_items}</ul>'
            f'</td></tr>'
        )
    else:
        attach_block = ""

    intro_block = (
        f'<p style="margin:0 0 12px;font-size:12px;line-height:1.6;color:#333;">'
        f'{custom_intro}</p>'
        if custom_intro else ""
    )

    kpi_section = (
        f'<!-- KPI Summary Tiles -->'
        f'<tr><td style="padding:4px 28px 2px;">'
        f'<p style="margin:0 0 6px;font-size:12px;font-weight:700;color:{primary};">'
        f'&#128202; Zone KPI Summary — {as_of}</p>'
        f'{kpi_tiles_html}'
        f'</td></tr>'
    ) if kpi_tiles_html else ""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#F0F2F6;
             font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0"
       style="background:#F0F2F6;padding:20px 0;">
  <tr><td align="center">
  <table width="700" cellpadding="0" cellspacing="0"
         style="background:#fff;border-radius:12px;overflow:hidden;
                box-shadow:0 4px 20px rgba(0,48,135,0.12);">

    <!-- Logo strip (white background) -->
    <tr><td style="background:#ffffff;padding:12px 24px;border-bottom:2px solid #e8eef8;">
      <table width="100%" cellpadding="0" cellspacing="0"><tr>
        <td width="64" style="vertical-align:middle;">
          {'<img src="' + _logo_uri + '" alt="HPCL" width="56" height="44" style="display:block;width:56px;height:44px;max-width:56px;max-height:44px;object-fit:contain;" />' if _logo_uri else '<div style="font-size:20px;font-weight:900;color:#003087;letter-spacing:.04em;">HPCL</div>'}
        </td>
        <td style="vertical-align:middle;padding-left:14px;">
          <div style="font-size:13px;font-weight:700;color:{accent};letter-spacing:0.02em;line-height:1.3;">
            Hindustan Petroleum Corporation Limited
          </div>
          <div style="font-size:10px;color:#555;margin-top:2px;">
            Supply, Operations &amp; Distribution
          </div>
        </td>
        <td align="right" style="vertical-align:middle;">
          <span style="font-size:11px;color:#666;">Data as of: <b style="color:{primary};">{as_of}</b></span>
        </td>
      </tr></table>
    </td></tr>

    <!-- Alert title strip (dark blue gradient) -->
    <tr><td style="background:linear-gradient(135deg,{primary} 0%,{secondary} 100%);
                   padding:14px 24px;">
      <div style="font-size:18px;font-weight:900;color:#fff;letter-spacing:0.04em;">
        &#128204;&nbsp; HPCL — SOD Exception Alert
      </div>
    </td></tr>

    <!-- Zone pill -->
    <tr><td style="padding:14px 28px 6px;">
      <div style="display:inline-block;background:#E8F0FE;
                  border-left:4px solid {primary};border-radius:6px;
                  padding:8px 16px;">
        <span style="font-size:14px;font-weight:700;color:{primary};">
          &#128205;&nbsp;{zone_name}
        </span>
        &nbsp;
        <span style="font-size:11px;color:#555;">
          {total_locations} location(s) with open exceptions
        </span>
      </div>
    </td></tr>

    <!-- Intro para -->
    <tr><td style="padding:8px 28px 10px;">
      {intro_block}
      <p style="margin:0;font-size:12px;color:#444;line-height:1.6;">
        Please find below a summary of <b>open SOD exceptions</b> for your zone
        as on <b>{as_of}</b>. Detailed data is attached as Excel file(s).
        Kindly take necessary action to clear the pending items at the earliest.
      </p>
    </td></tr>

    {kpi_section}

    <!-- Location-wise exception table heading -->
    <tr><td style="padding:14px 28px 4px;">
      <p style="margin:0;font-size:12px;font-weight:700;color:{primary};">
        &#127981; Location-wise Exception Detail
      </p>
    </td></tr>

    <!-- Exception summary table -->
    <tr><td style="padding:0 28px 16px;">
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border-collapse:collapse;border-radius:8px;overflow:hidden;
                    border:1px solid #D5E2F3;">
        <thead>{table_header}</thead>
        <tbody>{rows_html}</tbody>
      </table>
    </td></tr>

    {attach_block}

    <!-- Footer -->
    <tr><td style="background:#F4F8FF;padding:14px 28px;
                   border-top:1px solid #D5E2F3;">
      <p style="margin:0;font-size:10px;color:#888;line-height:1.6;">
        This is an auto-generated alert from the
        <b>HPCL SOD Exception Dashboard (LIVEDB)</b>.
        For queries contact
        <a href="mailto:{SENDER_EMAIL}"
           style="color:{primary};">{SENDER_EMAIL}</a>.
      </p>
    </td></tr>

  </table>
  </td></tr>
</table>
</body></html>"""


# ── Consolidated mail (all-zones HQ report) ───────────────────────────────────

#: Primary recipient for the consolidated HQ report
CONSOLIDATED_TO: str  = "shoaibrehman@hpcl.in"
CONSOLIDATED_CC: str  = "SOD.OPNS.HQO@hpcl.in"
CONSOLIDATED_BCC: str = "bhsgk@hpcl.in; shubham.tayal@hpcl.in"


def build_consolidated_excel_attachment(
    zone_exception_summary_df: pd.DataFrame,
    all_exception_plant_df: pd.DataFrame,
    detail_dfs: Dict[str, pd.DataFrame],
    selected_exceptions: List[str],
    as_of_date: str = "",
) -> List[Tuple[str, bytes]]:
    """Build one consolidated Excel workbook (all zones).

    Sheets:
      • Zone Summary       — zone-wise totals
      • Location Detail    — all-zone location-wise exceptions
      • <ExceptionLabel>   — raw detail per selected exception type (all zones)

    Returns [(filename, bytes)] — a single-element list.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    NAVY   = PatternFill("solid", fgColor="003087")
    BLUE   = PatternFill("solid", fgColor="0057A8")
    WHITE  = Font(bold=True, color="FFFFFF", size=10)
    thin   = Side(style="thin", color="D5E2F3")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft    = Alignment(horizontal="left",   vertical="center")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        def _style_sheet(ws, hdr_fill):
            for cell in ws[1]:
                cell.fill      = hdr_fill
                cell.font      = WHITE
                cell.alignment = ctr
                cell.border    = brd
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = lft
                    cell.border    = brd
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 42)
            ws.row_dimensions[1].height = 28

        # Sheet 1: Zone Summary
        if zone_exception_summary_df is not None and not zone_exception_summary_df.empty:
            zone_exception_summary_df.to_excel(
                writer, index=False, sheet_name="Zone Summary"
            )
            _style_sheet(writer.sheets["Zone Summary"], NAVY)

        # Sheet 2: Location Detail
        if all_exception_plant_df is not None and not all_exception_plant_df.empty:
            loc_df = all_exception_plant_df.copy()
            drop = [c for c in loc_df.columns if c.startswith("_")]
            loc_df = loc_df.drop(columns=drop, errors="ignore")
            loc_df.to_excel(writer, index=False, sheet_name="Location Detail")
            _style_sheet(writer.sheets["Location Detail"], BLUE)

        # One sheet per exception type (all zones)
        for exc_key in selected_exceptions:
            label     = EXCEPTION_LABELS.get(exc_key, exc_key)
            detail_df = detail_dfs.get(exc_key)
            if detail_df is None or detail_df.empty:
                continue
            df_out = detail_df.copy()
            drop   = [c for c in df_out.columns
                      if c.startswith("_") or c in {"index", "level_0"}]
            df_out = df_out.drop(columns=drop, errors="ignore").reset_index(drop=True)
            sheet_name = label[:31]
            df_out.to_excel(writer, index=False, sheet_name=sheet_name)
            _style_sheet(writer.sheets[sheet_name], BLUE)

    date_tag  = (as_of_date or datetime.now().strftime("%d%b%Y")).replace(" ", "")
    filename  = f"HPCL_SOD_Consolidated_{date_tag}.xlsx"
    return [(filename, buf.getvalue())]


def build_consolidated_email_html(
    as_of_date: str,
    zone_exception_summary_df: pd.DataFrame,
    all_exception_plant_df: pd.DataFrame,
    selected_exceptions: List[str],
    grand_kpis: Optional[dict] = None,
    custom_intro: str = "",
    attachment_names: Optional[List[str]] = None,
) -> str:
    """Build professional HTML email body for the consolidated HQ report (all zones)."""

    # ── Palette ───────────────────────────────────────────────────────────────
    C_HDR       = "#1C2533"    # dark charcoal — page header / title bar
    C_HDR_MID   = "#2E3F55"    # medium dark for sub-elements
    C_TBL_HDR   = "#1A4775"    # deep navy-blue for table column headers
    C_TBL_HDR_B = "#153760"    # table header border
    C_GOLD_BDR  = "#C9A227"    # gold underline accent on letterhead
    C_BODY_BG   = "#ECEEF1"    # outer page background
    C_WHITE     = "#FFFFFF"
    C_BORDER    = "#D0D5DF"    # table/cell borders
    C_ZONE_HDR  = "#DFE8F2"    # zone group header row background
    C_ALT       = "#F5F7FA"    # alternating row
    C_RED       = "#B91C1C"    # exception value > 0
    C_GREEN     = "#166534"    # zero / compliant
    C_TEXT      = "#1A202C"    # primary body text
    C_MUTED     = "#5C6878"    # secondary / label text

    _logo_uri = _logo_data_uri()
    as_of     = as_of_date or datetime.now().strftime("%d %b %Y")
    kpis      = grand_kpis or {}

    # Ordinal day suffix: "31" → "31st", "1" → "1st", etc.
    def _ordinal_date(date_str: str) -> str:
        try:
            _d = datetime.strptime(date_str.strip(), "%d %b %Y")
            day = _d.day
            sfx = ("th" if 11 <= day <= 13
                   else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th"))
            return f"{day}<sup>{sfx}</sup> {_d.strftime('%B %Y')}"
        except Exception:
            return date_str

    as_of_display = _ordinal_date(as_of)

    def _n(key, default=0):  return kpis.get(key, default)
    def _fmt_n(v):
        try:    return f"{int(v):,}"
        except: return str(v)
    def _fmt_l(v):
        try:
            n = float(v)
            if n >= 1_000_000: return f"{n/1_000_000:.2f}M"
            if n >= 1_000:     return f"{n/1_000:.1f}K"
            return f"{n:,.0f}"
        except: return str(v)

    # ── Logo (constrained in a fixed-width td so Outlook can't stretch it) ────
    if _logo_uri:
        logo_html = (
            f'<img src="{_logo_uri}" alt="HPCL" '
            f'width="56" height="44" '
            f'style="display:block;width:56px;height:44px;'
            f'max-width:56px;max-height:44px;object-fit:contain;" />'
        )
    else:
        logo_html = (
            f'<span style="font-size:20px;font-weight:900;'
            f'color:{C_GOLD_BDR};letter-spacing:2px;">HPCL</span>'
        )

    # ── KPI cards (filled bg + rounded corners) ───────────────────────────────
    def _kpi_card(label, value, unit, accent, bg):
        return (
            f'<td style="width:25%;padding:4px 5px;">'
            f'<div style="background:{bg};'
            f'border:1px solid {accent};border-radius:7px;'
            f'padding:10px 13px;">'
            f'<div style="font-size:9px;font-weight:700;color:{accent};'
            f'text-transform:uppercase;letter-spacing:0.4px;'
            f'margin-bottom:4px;">{label}</div>'
            f'<div style="font-size:19px;font-weight:700;color:{C_TEXT};'
            f'font-variant-numeric:tabular-nums;line-height:1.1;">{value}</div>'
            f'<div style="font-size:8.5px;color:{C_MUTED};margin-top:2px;">{unit}</div>'
            f'</div></td>'
        )

    kpi_row1 = (
        '<table width="100%" cellpadding="0" cellspacing="0"><tr>'
        + _kpi_card("Pending DCs",       _fmt_n(_n("Pending DC")),           "documents",  "#B91C1C", "#FEF2F2")
        + _kpi_card("Open Deliveries",   _fmt_n(_n("Open Delivery")),         "deliveries", "#B91C1C", "#FEF2F2")
        + _kpi_card("Open In-Transit",   _fmt_n(_n("Open In-Transit")),       "STO orders", "#1D4ED8", "#EFF6FF")
        + _kpi_card("Open Sales Orders", _fmt_n(_n("Open Sales Order")),      "sales docs", "#92400E", "#FFFBEB")
        + '</tr></table>'
    )
    kpi_row2 = (
        '<table width="100%" cellpadding="0" cellspacing="0"><tr>'
        + _kpi_card("Pending Invoices",  _fmt_n(_n("Pending Invoice")),       "deliveries", "#92400E", "#FFFBEB")
        + _kpi_card("Shortage Sales",    _fmt_l(_n("Shortage Sales (Ltrs)")), "Litres",     "#B91C1C", "#FEF2F2")
        + _kpi_card("Shortage STO",      _fmt_l(_n("Shortage STO (Ltrs)")),   "Litres",     "#B91C1C", "#FEF2F2")
        + _kpi_card("Total Exceptions",  _fmt_n(_n("Total Exceptions")),      "all types",  "#5B21B6", "#F5F3FF")
        + '</tr></table>'
    )

    # ── Column helpers ─────────────────────────────────────────────────────────
    col_labels_map = {
        **EXCEPTION_LABELS,
        "Shortage Sales (Ltrs)": "Shrt Sales (L)",
        "Shortage STO (Ltrs)":   "Shrt STO (L)",
    }
    qty_cols_set = {"Shortage Sales (Ltrs)", "Shortage STO (Ltrs)"}

    exc_cols_z = [c for c in selected_exceptions
                  if zone_exception_summary_df is not None
                  and c in zone_exception_summary_df.columns]
    qty_cols_z = [c for c in qty_cols_set
                  if zone_exception_summary_df is not None
                  and c in zone_exception_summary_df.columns]
    all_zone_cols = exc_cols_z + qty_cols_z

    # Count zones / locations
    n_zones = (
        int(zone_exception_summary_df["Zone Name"].nunique())
        if zone_exception_summary_df is not None and not zone_exception_summary_df.empty else 0
    )
    n_locs = (
        int(all_exception_plant_df["Plant Name"].nunique())
        if all_exception_plant_df is not None and not all_exception_plant_df.empty else 0
    )

    # ── Table header helper ────────────────────────────────────────────────────
    def _th(text, align="center", w=""):
        ws = f"width:{w};" if w else ""
        return (
            f'<th style="{ws}background:{C_TBL_HDR};color:{C_WHITE};'
            f'font-size:10px;font-weight:700;text-transform:uppercase;'
            f'letter-spacing:0.35px;padding:7px 8px;text-align:{align};'
            f'border:1px solid {C_TBL_HDR_B};white-space:nowrap;">{text}</th>'
        )

    # ── Zone summary table ─────────────────────────────────────────────────────
    zone_hdr_row = (
        f'<tr>{_th("Zone","left","150px")}{_th("Locs","center","38px")}'
        f'{_th("Total Exc.","center","50px")}'
        + "".join(_th(col_labels_map.get(c, c)) for c in all_zone_cols)
        + "</tr>"
    )

    zone_rows_html = ""
    if zone_exception_summary_df is not None and not zone_exception_summary_df.empty:
        for i, (_, row) in enumerate(
            zone_exception_summary_df.sort_values("Zone Name", ascending=True).iterrows()
        ):
            bg  = C_WHITE if i % 2 == 0 else C_ALT
            tdc = (f'style="padding:5px 8px;font-size:10px;text-align:center;'
                   f'border:1px solid {C_BORDER};background:{bg};"')
            tdl = (f'style="padding:5px 8px;font-size:10px;font-weight:600;'
                   f'color:{C_HDR};text-align:left;border:1px solid {C_BORDER};'
                   f'background:{bg};"')

            def _zcell(val, is_qty=False):
                try:
                    n = float(val or 0)
                    txt = f"{n:,.2f}" if is_qty else f"{int(n):,}"
                    clr = C_RED if n > 0 else C_GREEN
                    fw  = "700" if n > 0 else "400"
                    return f'<td {tdc}><span style="color:{clr};font-weight:{fw};">{txt}</span></td>'
                except Exception:
                    return f'<td {tdc}>{val}</td>'

            tot  = int(pd.to_numeric(row.get("Total Exceptions", 0), errors="coerce") or 0)
            locs = int(pd.to_numeric(row.get("Locations",         0), errors="coerce") or 0)
            zone_rows_html += (
                f"<tr><td {tdl}>{row.get('Zone Name','')}</td>"
                f'<td {tdc}>{locs}</td>'
                f'<td {tdc}><span style="color:{C_RED};font-weight:700;">{tot:,}</span></td>'
                + "".join(_zcell(row.get(c, 0), c in qty_cols_set) for c in all_zone_cols)
                + "</tr>"
            )
    else:
        n_col = len(all_zone_cols) + 3
        zone_rows_html = (
            f'<tr><td colspan="{n_col}" style="padding:12px;text-align:center;'
            f'color:{C_MUTED};font-size:10px;">No zone data available.</td></tr>'
        )

    # ── Location-wise detail (zone-grouped) ───────────────────────────────────
    loc_exc_cols = [c for c in (exc_cols_z + qty_cols_z)
                    if all_exception_plant_df is not None
                    and c in all_exception_plant_df.columns]

    loc_table_hdr = ""
    loc_detail_html = ""
    if all_exception_plant_df is not None and not all_exception_plant_df.empty:
        loc_df = all_exception_plant_df.copy()
        # Only show locations with at least one open exception
        if "Total Exceptions" in loc_df.columns:
            loc_df = loc_df[
                pd.to_numeric(loc_df["Total Exceptions"], errors="coerce").fillna(0) > 0
            ]
        # Sort: zone alphabetically, then descending by Total Exceptions
        sort_cols = ["Zone Name"] + (
            ["Total Exceptions"] if "Total Exceptions" in loc_df.columns else []
        )
        try:
            loc_df = loc_df.sort_values(
                sort_cols, ascending=[True] + [False] * (len(sort_cols) - 1)
            )
        except Exception:
            pass

        code_col = next(
            (c for c in ["Plant Code", "Location Code", "Plant"]
             if c in loc_df.columns), None
        )

        code_th  = _th("Code", "center", "48px") if code_col else ""
        loc_table_hdr = (
            f'<tr>{_th("Location Name","left","190px")}{code_th}'
            f'{_th("Total","center","42px")}'
            + "".join(_th(col_labels_map.get(c, c)) for c in loc_exc_cols)
            + "</tr>"
        )

        # Build zone dict for totals lookup
        zone_totals_map: dict = {}
        if zone_exception_summary_df is not None and not zone_exception_summary_df.empty:
            for _, zr in zone_exception_summary_df.iterrows():
                zn = str(zr.get("Zone Name", ""))
                zt = int(pd.to_numeric(zr.get("Total Exceptions", 0), errors="coerce") or 0)
                lc = int(pd.to_numeric(zr.get("Locations", 0), errors="coerce") or 0)
                zone_totals_map[zn] = (zt, lc)

        current_zone = None
        n_loc_cols   = 3 + len(loc_exc_cols) + (1 if code_col else 0)

        for _, row in loc_df.iterrows():
            zone = str(row.get("Zone Name", ""))
            if zone != current_zone:
                current_zone = zone
                zt, zlc = zone_totals_map.get(zone, (0, 0))
                loc_detail_html += (
                    f'<tr><td colspan="{n_loc_cols}" '
                    f'style="background:{C_ZONE_HDR};color:{C_HDR};'
                    f'font-size:9.5px;font-weight:700;text-transform:uppercase;'
                    f'letter-spacing:0.4px;padding:6px 8px;'
                    f'border:1px solid {C_BORDER};border-top:2px solid {C_TBL_HDR};">'
                    f'{zone}'
                    f'<span style="color:{C_MUTED};font-weight:400;'
                    f'text-transform:none;font-size:8.5px;padding-left:10px;">'
                    f'{zlc} location(s) &nbsp;|&nbsp; Total Exceptions: {zt:,}'
                    f'</span></td></tr>'
                )

            tdc_l = (f'style="padding:5px 8px;font-size:9.5px;font-weight:500;'
                     f'text-align:left;border:1px solid {C_BORDER};color:{C_TEXT};"')
            tdc_c = (f'style="padding:5px 8px;font-size:9.5px;text-align:center;'
                     f'border:1px solid {C_BORDER};color:{C_MUTED};"')
            tdc_n = (f'style="padding:5px 8px;font-size:9.5px;text-align:center;'
                     f'border:1px solid {C_BORDER};"')

            tot = float(pd.to_numeric(row.get("Total Exceptions", 0), errors="coerce") or 0)
            code_td = (
                f'<td {tdc_c}>{row.get(code_col, "")}</td>' if code_col else ""
            )
            loc_detail_html += (
                f'<tr><td {tdc_l}>{row.get("Plant Name", "")}</td>'
                f'{code_td}'
                f'<td {tdc_n}><span style="color:{C_RED if tot>0 else C_GREEN};'
                f'font-weight:{"700" if tot>0 else "400"};">{int(tot):,}</span></td>'
            )
            for c in loc_exc_cols:
                v = float(pd.to_numeric(row.get(c, 0), errors="coerce") or 0)
                is_qty = c in qty_cols_set
                txt = f"{v:,.2f}" if is_qty else f"{int(v):,}"
                clr = C_RED if v > 0 else C_GREEN
                fw  = "700" if v > 0 else "400"
                loc_detail_html += (
                    f'<td {tdc_n}><span style="color:{clr};font-weight:{fw};">'
                    f'{txt}</span></td>'
                )
            loc_detail_html += "</tr>"

    # ── Attachment badges ──────────────────────────────────────────────────────
    if attachment_names:
        badges = "".join(
            f'<span style="display:inline-block;background:#F0F4FF;'
            f'border:1px solid #C7D2FE;border-radius:3px;padding:2px 8px;'
            f'font-size:8.5px;margin:2px 3px 2px 0;">{n}</span>'
            for n in attachment_names
        )
        attach_block = (
            f'<tr><td style="padding:6px 20px 14px;">'
            f'<div style="font-size:8.5px;font-weight:700;color:{C_MUTED};'
            f'text-transform:uppercase;letter-spacing:0.4px;margin-bottom:4px;">'
            f'Attachments</div>{badges}</td></tr>'
        )
    else:
        attach_block = ""

    intro_block = (
        f'<tr><td style="padding:10px 20px 4px;">'
        f'<p style="margin:0;font-size:10px;line-height:1.6;color:{C_TEXT};">'
        f'{custom_intro}</p></td></tr>'
    ) if custom_intro else ""

    # ── Section divider helper ─────────────────────────────────────────────────
    def _section(title, subtitle=""):
        sub = (f'<span style="font-weight:400;text-transform:none;'
               f'letter-spacing:0;font-size:8px;">&nbsp; {subtitle}</span>'
               if subtitle else "")
        return (
            f'<tr><td style="padding:12px 20px 5px;">'
            f'<div style="font-size:8.5px;font-weight:700;color:{C_MUTED};'
            f'text-transform:uppercase;letter-spacing:0.6px;'
            f'border-bottom:1px solid {C_BORDER};padding-bottom:4px;">'
            f'{title}{sub}</div></td></tr>'
        )

    # ── Assemble ───────────────────────────────────────────────────────────────
    return f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:{C_BODY_BG};
             font-family:Arial,'Helvetica Neue',Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0"
       style="background:{C_BODY_BG};padding:16px 0;">
  <tr><td align="center">
  <table width="720" cellpadding="0" cellspacing="0"
         style="background:{C_WHITE};border:1px solid {C_BORDER};">

    <!-- Header -->
    <tr><td style="padding:10px 18px 10px;background:{C_WHITE};
                   border-bottom:3px solid {C_GOLD_BDR};">
      <table width="100%" cellpadding="0" cellspacing="0"><tr>
        <td width="64" style="vertical-align:middle;">{logo_html}</td>
        <td style="vertical-align:middle;padding-left:10px;">
          <div style="font-size:12px;font-weight:700;color:{C_HDR};
                      letter-spacing:0.2px;line-height:1.4;">
            HINDUSTAN PETROLEUM CORPORATION LIMITED
          </div>
          <div style="font-size:9.5px;color:{C_MUTED};margin-top:2px;">
            Supply, Operations &amp; Distribution &nbsp;&bull;&nbsp; SOD Exception Dashboard
          </div>
        </td>
        <td align="right" style="vertical-align:middle;white-space:nowrap;
                                  padding-left:12px;">
          <div style="font-size:8.5px;color:{C_MUTED};text-transform:uppercase;
                      letter-spacing:0.3px;">Data as of</div>
          <div style="font-size:14px;font-weight:700;color:{C_HDR};">{as_of_display}</div>
          <div style="font-size:7px;font-weight:700;letter-spacing:0.8px;
                      color:{C_WHITE};background:{C_HDR};
                      display:inline-block;padding:2px 6px;margin-top:2px;">
            INTERNAL USE ONLY
          </div>
        </td>
      </tr></table>
    </td></tr>

    <!-- Title bar -->
    <tr><td style="background:{C_HDR};padding:9px 18px;">
      <table width="100%" cellpadding="0" cellspacing="0"><tr>
        <td>
          <div style="font-size:14px;font-weight:700;color:{C_WHITE};
                      letter-spacing:0.15px;">
            SOD Exception Report &mdash; All Zones Consolidated
          </div>
        </td>
        <td align="right" style="white-space:nowrap;">
          <span style="font-size:8px;color:#94A3B8;">
            {n_zones} Zones &nbsp;&bull;&nbsp; {n_locs} Locations with exceptions
          </span>
        </td>
      </tr></table>
    </td></tr>

    {intro_block}

    {_section("Exception Summary — All India")}
    <tr><td style="padding:2px 16px 2px;">{kpi_row1}</td></tr>
    <tr><td style="padding:2px 16px 10px;">{kpi_row2}</td></tr>

    {_section("Zone-wise Exception Summary")}
    <tr><td style="padding:0 20px 10px;">
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border-collapse:collapse;border:1px solid {C_BORDER};">
        <thead>{zone_hdr_row}</thead>
        <tbody>{zone_rows_html}</tbody>
      </table>
    </td></tr>

    {_section("Location-wise Detail — Zone Consolidated",
              "showing only locations with open exceptions")}
    <tr><td style="padding:0 20px 10px;">
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border-collapse:collapse;border:1px solid {C_BORDER};">
        <thead>{loc_table_hdr}</thead>
        <tbody>{loc_detail_html if loc_detail_html
                else f'<tr><td colspan="10" style="padding:10px;text-align:center;'
                     f'color:{C_MUTED};font-size:9px;">No location data available.</td></tr>'
               }</tbody>
      </table>
    </td></tr>

    {attach_block}

    <!-- Footer -->
    <tr><td style="background:{C_ALT};padding:8px 18px;
                   border-top:1px solid {C_BORDER};">
      <p style="margin:0;font-size:8px;color:{C_MUTED};line-height:1.5;">
        Auto-generated by HPCL SOD Exception Dashboard (LIVEDB).
        Queries: <a href="mailto:{SENDER_EMAIL}"
                    style="color:{C_HDR};">{SENDER_EMAIL}</a>
        &nbsp;&bull;&nbsp; Confidential — Internal Use Only.
      </p>
    </td></tr>

  </table>
  </td></tr>
</table>
</body></html>"""


def send_consolidated_mail(
    all_exception_plant_df: pd.DataFrame,
    zone_exception_summary_df: pd.DataFrame,
    detail_dfs: Dict[str, pd.DataFrame],
    selected_exceptions: List[str],
    as_of_date: str = "",
    custom_intro: str = "",
    test_mode: bool = False,
    test_email: str = "",
    grand_kpis: Optional[dict] = None,
) -> dict:
    """Send one consolidated exception mail (all zones) to HQ / shoaibrehman@hpcl.in.

    test_mode=True  → mail goes to test_email only (no real recipients, no BCC).
    Returns {"ok": bool, "mode": str, "msg": str, "attachments": [filenames]}
    """
    avail, reason = outlook_available()
    if not avail:
        return {"ok": False, "msg": reason}

    as_of = as_of_date or datetime.now().strftime("%d %b %Y")

    # Build grand-total KPI dict from zone summary if not provided
    if grand_kpis is None and zone_exception_summary_df is not None and not zone_exception_summary_df.empty:
        grand_kpis = {}
        for col in zone_exception_summary_df.columns:
            if col not in {"Zone Name", "Locations"}:
                grand_kpis[col] = float(
                    pd.to_numeric(zone_exception_summary_df[col], errors="coerce").fillna(0).sum()
                )

    # Build consolidated Excel
    attachments = build_consolidated_excel_attachment(
        zone_exception_summary_df, all_exception_plant_df,
        detail_dfs, selected_exceptions, as_of_date=as_of,
    )
    attachment_names = [fname for fname, _ in attachments]

    # Build HTML body
    html_body = build_consolidated_email_html(
        as_of_date=as_of,
        zone_exception_summary_df=zone_exception_summary_df,
        all_exception_plant_df=all_exception_plant_df,
        selected_exceptions=selected_exceptions,
        grand_kpis=grand_kpis,
        custom_intro=custom_intro,
        attachment_names=attachment_names,
    )

    subject = f"[HPCL SOD Consolidated Exception Report] All Zones — {as_of}"
    if test_mode:
        subject  = f"[TEST] {subject}"
        to_str   = test_email or SENDER_EMAIL
        cc_str   = ""
        bcc_str  = ""
    else:
        to_str   = CONSOLIDATED_TO
        cc_str   = CONSOLIDATED_CC
        bcc_str  = CONSOLIDATED_BCC

    result = _send_via_outlook(
        to=to_str, subject=subject, html_body=html_body,
        cc=cc_str, bcc=bcc_str, attachments=attachments,
    )
    result["attachments"] = attachment_names
    return result


# ── Public API ────────────────────────────────────────────────────────────────

def send_exception_mail_for_zone(
    zone_name: str,
    all_exception_plant_df: pd.DataFrame,
    detail_dfs: Dict[str, pd.DataFrame],
    selected_exceptions: List[str],
    as_of_date: str = "",
    custom_intro: str = "",
    test_mode: bool = False,
    test_email: str = "",
    zone_kpi_dict: Optional[dict] = None,
) -> dict:
    """Send exception mail for one zone with per-exception Excel attachments.

    For each selected exception type that has data for the zone, one Excel file
    is built and attached: '<ZoneName>_<ExceptionLabel>.xlsx'.
    zone_kpi_dict: per-zone KPI values rendered as tiles in the email body.

    test_mode=True  → mail goes to test_email only (no zone recipients, no BCC).
    Returns {"ok": bool, "mode": str, "msg": str, "attachments": [filenames]}
    """
    avail, reason = outlook_available()
    if not avail:
        return {"ok": False, "msg": reason}

    contacts = ZONE_EMAIL_MAP.get(zone_name)
    if not contacts:
        return {"ok": False,
                "msg": f"No email contacts configured for zone: {zone_name}"}

    # Summary rows for this zone (for mail body)
    zone_summary = all_exception_plant_df[
        all_exception_plant_df["Zone Name"] == zone_name
    ].copy()

    # Build per-exception Excel attachments
    attachments = build_zone_excel_attachments(zone_name, detail_dfs, selected_exceptions)
    if not attachments:
        return {"ok": False,
                "msg": f"No exception data found for zone '{zone_name}' "
                       f"for the selected exception types."}

    attachment_names = [fname for fname, _ in attachments]

    html_body = build_exception_email_html(
        zone_name, as_of_date, zone_summary, selected_exceptions,
        custom_intro, attachment_names,
        zone_kpi_dict=zone_kpi_dict,
    )

    as_of   = as_of_date or datetime.now().strftime("%d %b %Y")
    subject = f"[HPCL SOD Exception Alert] {zone_name} — {as_of}"
    if test_mode:
        subject = f"[TEST] {subject}"

    if test_mode:
        to_str  = test_email or SENDER_EMAIL
        cc_str  = ""
        bcc_str = ""
    else:
        to_str  = contacts["to"]
        cc_str  = contacts.get("cc", "")
        bcc_str = "; ".join(BCC_EMAILS)

    result = _send_via_outlook(
        to=to_str,
        subject=subject,
        html_body=html_body,
        cc=cc_str,
        bcc=bcc_str,
        attachments=attachments,
    )
    result["attachments"] = attachment_names
    return result
